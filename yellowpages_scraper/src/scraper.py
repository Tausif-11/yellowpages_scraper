import os
import json
import time
import random
from bs4 import BeautifulSoup
import pandas as pd
from curl_cffi import requests

class YellowPagesScraper:
    def __init__(self, search_term, location, max_pages=20):
        self.base_url = "https://www.yellowpages.com/search"
        self.search_term = search_term
        self.location = location
        self.max_pages = max_pages
        self.results = []
        
        # Initialize a persistent browser session
        self.session = requests.Session()
        
        # Define baseline organic headers
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Connection": "keep-alive",
            "Sec-Ch-Ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1"
        }

    def fetch_page(self, page_num, current_url):
        """Fetches raw HTML using persistent sessions and progressive referer chaining."""
        params = {
            "search_terms": self.search_term,
            "geo_location_terms": self.location,
            "page": page_num
        }
        
        try:
            print(f"[+] Querying Page {page_num}...")
            
            # Fire the request through the tracking session with Chrome emulation
            response = self.session.get(
                self.base_url, 
                params=params, 
                headers=self.headers,
                impersonate="chrome", 
                timeout=30
            )
            
            if response.status_code == 200:
                # Update the Referer dynamically for the subsequent page request
                self.headers["Referer"] = response.url
                return response.text
            else:
                print(f"[-] Failed to fetch Page {page_num}. Status Code: {response.status_code}")
                return None
        except Exception as e:
            print(f"[!] Network connection exception on Page {page_num}: {e}")
            return None

    def parse_json_ld(self, html_content):
        """Parses internal schema-org dictionaries safely out of script elements."""
        soup = BeautifulSoup(html_content, 'html.parser')
        script_tags = soup.find_all('script', type='application/ld+json')
        
        page_listings_count = 0
        
        for tag in script_tags:
            try:
                if not tag.string:
                    continue
                data = json.loads(tag.string)
                
                if isinstance(data, list):
                    items = data
                elif isinstance(data, dict):
                    items = [data]
                else:
                    continue
                
                for item in items:
                    if item.get("@type") == "LocalBusiness":
                        address_info = item.get("address", {})
                        street = address_info.get("streetAddress", "N/A")
                        locality = address_info.get("addressLocality", "N/A")
                        region = address_info.get("addressRegion", "N/A")
                        zip_code = address_info.get("postalCode", "N/A")
                        
                        rating_info = item.get("aggregateRating", {})
                        rating = rating_info.get("ratingValue", "N/A") if rating_info else "N/A"
                        reviews = rating_info.get("reviewCount", "N/A") if rating_info else "N/A"
                        
                        business_record = {
                            "Business Name": item.get("name", "N/A"),
                            "Phone Number": item.get("telephone", "N/A"),
                            "Street Address": street,
                            "Locality": locality,
                            "State": region,
                            "Zip Code": zip_code,
                            "Rating": rating,
                            "Review Count": reviews,
                            "Operating Hours": ", ".join(item.get("openingHours", [])) if isinstance(item.get("openingHours"), list) else item.get("openingHours", "N/A"),
                            "Source URL": item.get("url", "N/A")
                        }
                        self.results.append(business_record)
                        page_listings_count += 1
            except (json.JSONDecodeError, TypeError, AttributeError):
                continue
                
        print(f"[√] Extracted {page_listings_count} valid business nodes from page payload.")
        return page_listings_count

    def run(self):
        """Orchestrates loop workflows and manages structural bypass sequences."""
        consecutive_failures = 0
        
        for page in range(1, self.max_pages + 1):
            # Calculate programmatic URL to feed initial Referer headers cleanly
            current_url = f"{self.base_url}?search_terms={self.search_term.replace(' ', '+')}&geo_location_terms={self.location.replace(' ', '+').replace(',', '%2C')}"
            if page > 1:
                current_url += f"&page={page-1}"
                
            html = self.fetch_page(page, current_url)
            
            if html:
                extracted = self.parse_json_ld(html)
                if extracted > 0:
                    consecutive_failures = 0
                else:
                    consecutive_failures += 1
            else:
                consecutive_failures += 1
                
            # Anti-compounding block layout strategy
            if consecutive_failures >= 3:
                print("[-] Aborting process loop due to multiple consecutive empty pages.")
                break
            
            # Keep delays dynamic and slightly larger to mimic organic browsing behavior
            if page < self.max_pages:
                delay = random.uniform(4.5, 8.0)
                print(f"[i] Dynamic throttling delay: Sleeping for {delay:.2f}s...")
                time.sleep(delay)
            
        self.save_all_formats()

    def save_all_formats(self):
        """Compiles scraped elements matrix into CSV, JSON, XML, HTML, and stylized Excel sheets."""
        if not self.results:
            print("[-] No analytical items discovered during script lifecycle.")
            return
            
        df = pd.DataFrame(self.results)
        df.drop_duplicates(subset=["Business Name", "Phone Number"], inplace=True)
        
        # Ensure we write out to the parent root directory correctly
        output_dir = os.path.join("..", "data") if os.path.basename(os.getcwd()) == "src" else "data"
        os.makedirs(output_dir, exist_ok=True)
        
        # 1. CSV Output
        csv_file = os.path.join(output_dir, "los_angeles_roofing_contractors.csv")
        df.to_csv(csv_file, index=False, encoding="utf-8")
        print(f"[Output] Flat CSV Compiled -> {csv_file}")
        
        # 2. JSON Output
        json_file = os.path.join(output_dir, "los_angeles_roofing_contractors.json")
        df.to_json(json_file, orient="records", indent=4, force_ascii=False)
        print(f"[Output] Schema JSON Matrix Compiled -> {json_file}")
        
        # 3. XML Output
        xml_file = os.path.join(output_dir, "los_angeles_roofing_contractors.xml")
        xml_content = ['<?xml version="1.0" encoding="utf-8"?>', '<BusinessListings>']
        for _, row in df.iterrows():
            xml_content.append('  <Listing>')
            for col, val in row.items():
                tag = col.replace(" ", "")
                clean_val = str(val).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                xml_content.append(f'    <{tag}>{clean_val}</{tag}>')
            xml_content.append('  </Listing>')
        xml_content.append('</BusinessListings>')
        with open(xml_file, "w", encoding="utf-8") as f:
            f.write("\n".join(xml_content))
        print(f"[Output] Structured Markup XML Compiled -> {xml_file}")
            
        # 4. HTML Output
        html_file = os.path.join(output_dir, "los_angeles_roofing_contractors.html")
        html_style = """<style>
            body { font-family: 'Segoe UI', Arial, sans-serif; background-color: #fafafa; margin: 30px; color: #222; }
            h2 { color: #1e3d59; border-bottom: 3px solid #ffc13b; padding-bottom: 10px; font-weight: 600; }
            table { width: 100%; border-collapse: collapse; margin-top: 20px; background: #fff; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }
            th { background-color: #1e3d59; color: #fff; text-align: left; padding: 12px; font-size: 14px; text-transform: uppercase; }
            td { padding: 12px; border-bottom: 1px solid #f1f1f1; font-size: 13px; }
            tr:nth-child(even) { background-color: #f8f9fa; }
            tr:hover { background-color: #f1f5f9; }
            a { color: #17a2b8; text-decoration: none; font-weight: 500; }
            a:hover { text-decoration: underline; }
        </style>"""
        with open(html_file, "w", encoding="utf-8") as f:
            f.write(f"<html><head><title>Scraped Directory Layout</title>{html_style}</head><body><h2>YellowPages Extract Table — Los Angeles Roofing Contractors</h2>{df.to_html(index=False)}</body></html>")
        print(f"[Output] Presentable Display HTML Table Compiled -> {html_file}")

        # 5. Excel Output via openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        excel_file = os.path.join(output_dir, "los_angeles_roofing_contractors.xlsx")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Contractor Directory', index=False)
            worksheet = writer.sheets['Contractor Directory']
            
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1E3D59', end_color='1E3D59', fill_type='solid')
            body_font = Font(name='Segoe UI', size=10)
            zebra_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
            link_font = Font(name='Segoe UI', size=10, color='17A2B8', underline='single')
            
            thin_border = Border(
                left=Side(style='thin', color='E5E9F0'), right=Side(style='thin', color='E5E9F0'),
                top=Side(style='thin', color='E5E9F0'), bottom=Side(style='thin', color='E5E9F0')
            )
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center' if col_num in [2,5,6,7,8] else 'left', vertical='center')
                
            for row_num in range(2, len(df) + 2):
                current_fill = zebra_fill if row_num % 2 == 0 else PatternFill(fill_type=None)
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = body_font if col_num != 10 else link_font
                    if current_fill.fill_type: 
                        cell.fill = current_fill
                    cell.border = thin_border
                    
                    if col_num in [2, 5, 6, 7, 8]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 42)
                
            worksheet.row_dimensions[1].height = 26
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
            
        print(f"[Output] Premium Styled Excel Workbook Compiled -> {excel_file}")
        print(f"\n[======= PIPELINE COMPLETE: {len(df)} UNIQUE LISTINGS COMPILED =======]\n")

if __name__ == "__main__":
    scraper = YellowPagesScraper(
        search_term="roofing contractors", 
        location="Los Angeles, CA", 
        max_pages=20
    )
    scraper.run()